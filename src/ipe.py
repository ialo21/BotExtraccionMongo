"""
Generación automática de IPE (Informe de Proceso de Extracción).

Abre la plantilla Excel, escribe datos en celdas específicas e inserta
capturas de pantalla en posiciones predefinidas sin alterar el formato.
"""
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


def generar_ipe(
    plantilla_path: Path,
    salida_path: Path,
    datos: dict[str, str],
    imagenes: list[Path],
    hoja: str = "Hoja 1",
    fila_base_img: int = 26,
    col_img: str = "D",
    espaciado_filas: int = 36,
    drive_url: str | None = None,
) -> Path:
    """
    Genera el archivo IPE a partir de una plantilla Excel.

    Args:
        plantilla_path:   Ruta al archivo Excel plantilla (no se modifica).
        salida_path:      Ruta donde se guardará el IPE generado.
        datos:            Diccionario celda -> valor. Ej: {"I3": "21/05/2024", "F4": "Auto002"}.
        imagenes:         Lista de rutas a archivos PNG/JPG (en orden cronológico).
        hoja:             Nombre de la hoja donde escribir (default: "Hoja 1").
        fila_base_img:    Fila inicial donde anclar la primera imagen (default: 24 = "Captura 1").
        col_img:          Columna de anclaje de las imágenes (default: "D").
        espaciado_filas:  Número de filas entre cada captura (default: 35).
        drive_url:        URL de la carpeta en Google Drive (se escribe en C134).

    Returns:
        Path al archivo IPE generado.

    Raises:
        FileNotFoundError: Si la plantilla no existe.
        ValueError: Si la hoja especificada no existe en la plantilla.
    """
    if not plantilla_path.exists():
        raise FileNotFoundError(f"Plantilla no encontrada: {plantilla_path}")

    print(f"[IPE] Cargando plantilla: {plantilla_path}")
    wb = load_workbook(plantilla_path)

    if hoja not in wb.sheetnames:
        raise ValueError(f"Hoja '{hoja}' no existe en la plantilla. Hojas disponibles: {wb.sheetnames}")

    ws = wb[hoja]

    # 1) Escribir datos en celdas específicas
    print(f"[IPE] Escribiendo {len(datos)} campos...")

    def _resolve_merged_cell(address: str):
        """Si la celda está dentro de un rango combinado, devuelve la esquina superior izquierda."""
        cell = ws[address]
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if address in merged_range:
                    return f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
        return address

    for celda, valor in datos.items():
        destino = _resolve_merged_cell(celda)
        ws[destino] = valor
    
    # Escribir URL de Drive en C134 si se proporciona
    if drive_url:
        destino_url = _resolve_merged_cell("C134")
        ws[destino_url] = drive_url
        print(f"[IPE] URL de Drive escrita en C134: {drive_url}")

    # 2) Insertar imágenes
    print(f"[IPE] Insertando {len(imagenes)} capturas...")
    fila = fila_base_img
    for idx, img_path in enumerate(imagenes, start=1):
        if not img_path.exists():
            print(f"  [aviso] Imagen no encontrada: {img_path.name}, saltando...")
            continue

        pic = XLImage(str(img_path))
        
        # Opcional: redimensionar si las capturas son muy grandes (ajusta según necesites)
        # Ejemplo: limitar ancho a 800px manteniendo proporción
        max_width = 800
        if pic.width > max_width:
            ratio = max_width / pic.width
            pic.width = max_width
            pic.height = int(pic.height * ratio)

        anchor = f"{col_img}{fila}"
        ws.add_image(pic, anchor)
        print(f"  → Captura {idx}: {img_path.name} en {anchor}")
        
        fila += espaciado_filas

    # 3) Guardar archivo resultante
    salida_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida_path)
    print(f"[IPE] ✓ Archivo generado: {salida_path}")
    
    return salida_path
